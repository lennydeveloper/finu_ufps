import sys
sys.path.append('.')

from openpyxl import load_workbook
from models import Proyecto

def get_projects_from_excel():
    projects = []
    data_file = r'C:\Users\Lenny\Documents\RESUMEN_FINU_2002-2022.xlsx'

    wb = load_workbook(data_file)
    ws = wb['Hoja3']

    for row in ws.iter_rows(min_row=2, values_only=True):
        project_item = Proyecto(
            contrato = row[0],
            anio = row[1],
            convocatoria = row[2],
            proyecto = row[3],
            tipo_proyecto = row[4],
            alianza = row[5],
            investigador = row[6],
            tipo_investigador = row[7],
            dedicacion_horas = row[8],
            grupo_investigacion = row[9],
            facultad = row[10],
            departamento = row[11],
            celular = row[12],
            supervisado = row[13],
            monto_financiado_finu = row[14],
            duracion = row[15],
            prorroga_1 = row[16],
            prorroga_2 = row[17],
            fecha_inicio_contrato = row[18],
            fecha_terminacion_contrato = row[19],
            informes_parciales_entregados = row[20],
            informe_final = row[21],
            presupuesto = row[22],
            presupuesto_no_aprobado = row[23],
            evaluador = row[24],
            estado = row[25],
            observaciones = row[26],
            producto = row[27],
            nombre_producto = row[28]
        )

        projects.append(project_item)

    return projects